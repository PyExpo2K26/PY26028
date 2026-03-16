"""
PyXcel — Pivot Engine
Generates pivot tables from Excel sheets using pandas.
"""
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter


def get_numeric_columns(filepath: str, sheet: str) -> list:
    try:
        df = pd.read_excel(filepath, sheet_name=sheet)
        return [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])]
    except Exception:
        return []


def get_all_columns(filepath: str, sheet: str) -> list:
    try:
        df = pd.read_excel(filepath, sheet_name=sheet)
        return list(df.columns)
    except Exception:
        return []


def generate_pivot(filepath, sheet, index_col, value_col,
                   agg_func="sum", columns_col=None, output_sheet="Pivot Table"):
    try:
        df    = pd.read_excel(filepath, sheet_name=sheet)
        pivot = pd.pivot_table(
            df, index=index_col, values=value_col,
            columns=columns_col if columns_col else None,
            aggfunc=agg_func, fill_value=0
        ).round(2)

        wb = openpyxl.load_workbook(filepath)
        if output_sheet in wb.sheetnames:
            del wb[output_sheet]
        ws = wb.create_sheet(output_sheet)

        for row in dataframe_to_rows(pivot, index=True, header=True):
            ws.append(row)

        for cell in ws[1]:
            cell.fill      = PatternFill("solid", fgColor="1e2035")
            cell.font      = Font(bold=True, color="7c83ff")
            cell.alignment = Alignment(horizontal="center")

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 30)

        wb.save(filepath)
        wb.close()

        return {
            "status": "success", "shape": pivot.shape,
            "rows": pivot.shape[0], "cols": pivot.shape[1],
            "output_sheet": output_sheet,
            "preview": pivot.head(5).to_string(),
            "message": f"Pivot table created in sheet '{output_sheet}'"
        }
    except Exception as e:
        return {"status": "error", "message": str(e)}


def generate_ai_pivot(filepath, sheet, instruction, output_sheet="AI Pivot"):
    try:
        from core.ollama_client      import ask_llama
        from core.workbook_inspector import get_context_string
        import re

        ctx = get_context_string(filepath)
        df  = pd.read_excel(filepath, sheet_name=sheet)

        system = (
            "You are a pandas expert. Respond ONLY with Python code — no markdown. "
            "DataFrame is `df`, pandas is `pd`. "
            "Create a pivot table and store it in variable `result_df`. "
            "Use pd.pivot_table(). Round numeric values to 2 decimal places. fill_value=0."
        )
        code = ask_llama(system, f"{ctx}\n\nColumns: {list(df.columns)}\nInstruction: {instruction}\n\nCreate pivot in result_df.")
        code = re.sub(r"```python|```", "", code).strip()

        local = {"df": df, "pd": pd}
        exec(code, local)
        result_df = local.get("result_df", None)

        if result_df is None:
            return {"status": "error", "message": "LLaMA did not create result_df"}

        wb = openpyxl.load_workbook(filepath)
        if output_sheet in wb.sheetnames:
            del wb[output_sheet]
        ws = wb.create_sheet(output_sheet)

        for row in dataframe_to_rows(result_df, index=True, header=True):
            ws.append(row)

        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor="1e2035")
            cell.font = Font(bold=True, color="7c83ff")

        wb.save(filepath)
        wb.close()

        return {
            "status": "success", "shape": result_df.shape,
            "output_sheet": output_sheet,
            "preview": result_df.head(5).to_string(),
            "code": code,
            "message": f"AI Pivot created in '{output_sheet}'"
        }
    except Exception as e:
        return {"status": "error", "message": str(e)}
