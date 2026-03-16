"""
Agent Worker — QThread workers for all AI operations.
"""
from PySide6.QtCore import QThread, Signal


class BaseWorker(QThread):
    result  = Signal(dict)
    error   = Signal(str)
    status  = Signal(str)

    def __init__(self, parent=None):
        super().__init__(parent)

    def run(self):
        raise NotImplementedError


class MacroWorker(BaseWorker):
    def __init__(self, filepath: str, instruction: str, parent=None):
        super().__init__(parent)
        self.filepath    = filepath
        self.instruction = instruction

    def run(self):
        try:
            from core.ollama_client      import ask_llama
            from core.workbook_inspector import get_context_string
            from core.code_executor      import execute_macro_code

            self.status.emit("Inspecting workbook structure...")
            ctx = get_context_string(self.filepath)
            self.status.emit("Generating code with LLaMA...")
            system = (
                "You are a Python/openpyxl expert. "
                "Respond ONLY with Python code — no markdown, no explanation, no backticks. "
                "Available variables: wb (openpyxl workbook), filepath (str). "
                "Imports available: openpyxl, PatternFill, Font, Alignment, "
                "Border, Side, get_column_letter, pd (pandas). "
                "Always end with wb.save(filepath)."
            )
            code = ask_llama(system, f"{ctx}\n\nUser instruction: {self.instruction}")
            self.status.emit("Executing generated code...")
            result = execute_macro_code(self.filepath, code)
            self.result.emit(result)
        except Exception as e:
            self.error.emit(str(e))


class FormulaWorker(BaseWorker):
    def __init__(self, description: str, context: str = "", parent=None):
        super().__init__(parent)
        self.description = description
        self.context     = context

    def run(self):
        try:
            from core.ollama_client import ask_llama
            self.status.emit("Generating formula with LLaMA...")
            system = (
                "You are an Excel formula expert. "
                "Respond ONLY with the Excel formula starting with =. "
                "No explanation, no markdown, no extra text. "
                "Use modern functions like XLOOKUP, FILTER, LET, UNIQUE, SUMPRODUCT where appropriate."
            )
            formula = ask_llama(system, f"{self.context}\n\nFormula needed: {self.description}").strip()
            self.result.emit({"status": "success", "formula": formula})
        except Exception as e:
            self.error.emit(str(e))


class CleanerWorker(BaseWorker):
    def __init__(self, filepath: str, sheet: str, instructions: str, parent=None):
        super().__init__(parent)
        self.filepath     = filepath
        self.sheet        = sheet
        self.instructions = instructions

    def run(self):
        try:
            import pandas as pd
            from core.ollama_client import ask_llama
            from core.code_executor import execute_cleaning_code

            self.status.emit("Loading sheet data...")
            df  = pd.read_excel(self.filepath, sheet_name=self.sheet)
            ctx = (
                f"DataFrame columns: {list(df.columns)}\n"
                f"Shape: {df.shape}\n"
                f"Dtypes:\n{df.dtypes.to_string()}\n"
                f"Sample (first 5 rows):\n{df.head(5).to_string()}"
            )
            self.status.emit("Generating cleaning code with LLaMA...")
            system = (
                "You are a pandas expert. "
                "Respond ONLY with Python code — no markdown, no explanation. "
                "The DataFrame is available as `df`. pandas is imported as `pd`. "
                "Store the final result back in `df`."
            )
            code   = ask_llama(system, f"{ctx}\n\nCleaning instructions: {self.instructions}")
            self.status.emit("Applying transformations...")
            result = execute_cleaning_code(self.filepath, self.sheet, code)
            self.result.emit(result)
        except Exception as e:
            self.error.emit(str(e))


class ChatWorker(BaseWorker):
    def __init__(self, filepath: str, message: str, history: list, parent=None):
        super().__init__(parent)
        self.filepath = filepath
        self.message  = message
        self.history  = history

    def run(self):
        try:
            from core.ollama_client      import ask_llama
            from core.workbook_inspector import get_context_string

            self.status.emit("Reading workbook context...")
            ctx = get_context_string(self.filepath)
            self.status.emit("Thinking...")
            system = (
                "You are a helpful data analyst. "
                "Answer questions about the spreadsheet data clearly and concisely. "
                f"\n\n{ctx}"
            )
            response = ask_llama(system, self.message, history=self.history)
            self.result.emit({"status": "success", "response": response, "message": self.message})
        except Exception as e:
            self.error.emit(str(e))


class KpiWorker(BaseWorker):
    def __init__(self, filepath: str, sheet: str, parent=None):
        super().__init__(parent)
        self.filepath = filepath
        self.sheet    = sheet

    def run(self):
        try:
            import pandas as pd
            import json
            import re
            from core.ollama_client import ask_llama

            self.status.emit("Loading sheet data...")
            df  = pd.read_excel(self.filepath, sheet_name=self.sheet)
            ctx = (
                f"DataFrame columns: {list(df.columns)}\n"
                f"Shape: {df.shape}\n"
                f"Dtypes:\n{df.dtypes.to_string()}\n"
                f"Sample:\n{df.head(5).to_string()}"
            )
            self.status.emit("Identifying KPIs with LLaMA...")
            system = (
                "You are a business analyst. "
                "Respond ONLY with a valid JSON array — no markdown, no explanation. "
                'Each item must have exactly these keys: '
                '{"title": string, "value": string, "description": string, "trend": "up"|"down"|"neutral"}. '
                "Identify 4-6 relevant business KPIs from the data."
            )
            raw_kpis = ask_llama(system, ctx)
            self.status.emit("Computing KPI values...")
            code_system = (
                "You are a pandas expert. "
                "Respond ONLY with Python code — no markdown, no explanation. "
                "DataFrame is `df`, pandas is `pd`. "
                "Store computed KPI values in a dict called `kpi_values` where keys match the KPI titles exactly."
            )
            code = ask_llama(code_system, ctx)
            code = re.sub(r"```python|```", "", code).strip()
            local = {"df": df, "pd": pd}
            try:
                exec(code, local)
                computed = local.get("kpi_values", {})
            except Exception:
                computed = {}
            try:
                clean = re.sub(r"```json|```", "", raw_kpis).strip()
                kpis  = json.loads(clean)
                for k in kpis:
                    if k.get("title") in computed:
                        k["value"] = str(computed[k["title"]])
            except Exception:
                kpis = [{"title": "Parse Error", "value": "—", "description": raw_kpis[:120], "trend": "neutral"}]
            self.result.emit({"status": "success", "kpis": kpis})
        except Exception as e:
            self.error.emit(str(e))


class InspectorWorker(BaseWorker):
    def __init__(self, filepath: str, parent=None):
        super().__init__(parent)
        self.filepath = filepath

    def run(self):
        try:
            from core.workbook_inspector import inspect_workbook
            self.status.emit("Reading workbook...")
            data = inspect_workbook(self.filepath)
            self.result.emit({"status": "success", "data": data})
        except Exception as e:
            self.error.emit(str(e))
# ── Pivot Worker ─────────────────────────────────────────────
class PivotWorker(BaseWorker):
    def __init__(
        self, filepath: str, sheet: str,
        mode: str, instruction: str = "",
        index_col: str = "", value_col: str = "",
        agg_func: str = "sum", columns_col: str = None,
        parent=None
    ):
        super().__init__(parent)
        self.filepath    = filepath
        self.sheet       = sheet
        self.mode        = mode        # "manual" or "ai"
        self.instruction = instruction
        self.index_col   = index_col
        self.value_col   = value_col
        self.agg_func    = agg_func
        self.columns_col = columns_col

    def run(self):
        try:
            if self.mode == "ai":
                from core.pivot_engine import generate_ai_pivot
                self.status.emit("Generating AI pivot table...")
                result = generate_ai_pivot(
                    self.filepath, self.sheet,
                    self.instruction
                )
            else:
                from core.pivot_engine import generate_pivot
                self.status.emit("Building pivot table...")
                result = generate_pivot(
                    self.filepath, self.sheet,
                    self.index_col, self.value_col,
                    self.agg_func, self.columns_col
                )
            self.result.emit(result)
        except Exception as e:
            self.error.emit(str(e))


# ── Chart Worker ─────────────────────────────────────────────
class ChartWorker(BaseWorker):
    def __init__(
        self, filepath: str, sheet: str,
        chart_type: str, x_col: str, y_col: str,
        title: str = "", parent=None
    ):
        super().__init__(parent)
        self.filepath   = filepath
        self.sheet      = sheet
        self.chart_type = chart_type
        self.x_col      = x_col
        self.y_col      = y_col
        self.title      = title

    def run(self):
        try:
            from core.chart_engine import generate_chart
            self.status.emit(f"Generating {self.chart_type} chart...")
            result = generate_chart(
                self.filepath, self.sheet,
                self.chart_type, self.x_col,
                self.y_col, self.title
            )
            self.result.emit(result)
        except Exception as e:
            self.error.emit(str(e))


class ChartPreviewWorker(BaseWorker):
    """Generates chart preview bytes for display in GUI."""
    preview_ready = __import__('PySide6.QtCore', fromlist=['Signal']).Signal(bytes)

    def __init__(
        self, filepath: str, sheet: str,
        chart_type: str, x_col: str,
        y_col: str, title: str = "", parent=None
    ):
        super().__init__(parent)
        self.filepath   = filepath
        self.sheet      = sheet
        self.chart_type = chart_type
        self.x_col      = x_col
        self.y_col      = y_col
        self.title      = title

    def run(self):
        try:
            from core.chart_engine import get_chart_preview
            self.status.emit("Generating preview...")
            img_bytes = get_chart_preview(
                self.filepath, self.sheet,
                self.chart_type, self.x_col,
                self.y_col, self.title
            )
            if img_bytes:
                self.preview_ready.emit(img_bytes)
                self.result.emit({"status": "success"})
            else:
                self.result.emit({
                    "status" : "error",
                    "message": "Preview generation failed"
                })
        except Exception as e:
            self.error.emit(str(e))


# ── PDF Worker ───────────────────────────────────────────────
class PdfWorker(BaseWorker):
    def __init__(
        self, filepath: str, sheet: str,
        output_path: str, title: str = "",
        export_all: bool = False,
        include_summary: bool = True,
        parent=None
    ):
        super().__init__(parent)
        self.filepath        = filepath
        self.sheet           = sheet
        self.output_path     = output_path
        self.title           = title
        self.export_all      = export_all
        self.include_summary = include_summary

    def run(self):
        try:
            if self.export_all:
                from core.pdf_engine import export_all_sheets_to_pdf
                self.status.emit("Exporting all sheets to PDF...")
                result = export_all_sheets_to_pdf(
                    self.filepath,
                    self.output_path,
                    self.title
                )
            else:
                from core.pdf_engine import export_sheet_to_pdf
                self.status.emit("Exporting sheet to PDF...")
                result = export_sheet_to_pdf(
                    self.filepath, self.sheet,
                    self.output_path, self.title,
                    self.include_summary
                )
            self.result.emit(result)
        except Exception as e:
            self.error.emit(str(e))


# ── Merger Worker ────────────────────────────────────────────
class MergerWorker(BaseWorker):
    def __init__(
        self, files: list, output_path: str,
        mode: str = "sheets", parent=None
    ):
        super().__init__(parent)
        self.files       = files      # list of file paths
        self.output_path = output_path
        self.mode        = mode       # "sheets" or "rows"

    def run(self):
        try:
            import pandas as pd
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment

            if self.mode == "sheets":
                # Each file becomes a sheet
                self.status.emit("Merging files as separate sheets...")
                wb = openpyxl.Workbook()
                wb.remove(wb.active)

                for i, fpath in enumerate(self.files):
                    self.status.emit(
                        f"Processing file {i+1}/{len(self.files)}..."
                    )
                    try:
                        xl   = pd.ExcelFile(fpath)
                        base = __import__('os').path.splitext(
                            __import__('os').path.basename(fpath)
                        )[0][:25]

                        for sheet_name in xl.sheet_names:
                            df = xl.parse(sheet_name)
                            ws_name = f"{base}_{sheet_name}"[:31]

                            # Avoid duplicate sheet names
                            counter = 1
                            orig    = ws_name
                            while ws_name in wb.sheetnames:
                                ws_name = f"{orig[:27]}_{counter}"
                                counter += 1

                            ws = wb.create_sheet(ws_name)
                            from openpyxl.utils.dataframe import dataframe_to_rows
                            for row in dataframe_to_rows(
                                df, index=False, header=True
                            ):
                                ws.append(row)

                            # Style header
                            for cell in ws[1]:
                                cell.fill = PatternFill(
                                    "solid", fgColor="1e2035"
                                )
                                cell.font = Font(
                                    bold=True, color="7c83ff"
                                )
                    except Exception as fe:
                        continue

                wb.save(self.output_path)
                wb.close()

                self.result.emit({
                    "status"     : "success",
                    "mode"       : "sheets",
                    "files"      : len(self.files),
                    "output_path": self.output_path,
                    "message"    : f"Merged {len(self.files)} files into separate sheets"
                })

            elif self.mode == "rows":
                # Stack all files vertically
                self.status.emit("Merging files by stacking rows...")
                dfs = []
                for i, fpath in enumerate(self.files):
                    self.status.emit(
                        f"Reading file {i+1}/{len(self.files)}..."
                    )
                    try:
                        df       = pd.read_excel(fpath)
                        df["_source_file"] = __import__('os').path.basename(fpath)
                        dfs.append(df)
                    except Exception:
                        continue

                if not dfs:
                    self.result.emit({
                        "status" : "error",
                        "message": "No files could be read"
                    })
                    return

                merged = pd.concat(dfs, ignore_index=True)
                self.status.emit("Writing merged file...")

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Merged Data"

                from openpyxl.utils.dataframe import dataframe_to_rows
                for row in dataframe_to_rows(
                    merged, index=False, header=True
                ):
                    ws.append(row)

                # Style header
                for cell in ws[1]:
                    cell.fill = PatternFill("solid", fgColor="1e2035")
                    cell.font = Font(bold=True, color="7c83ff")

                wb.save(self.output_path)
                wb.close()

                self.result.emit({
                    "status"     : "success",
                    "mode"       : "rows",
                    "files"      : len(self.files),
                    "total_rows" : len(merged),
                    "output_path": self.output_path,
                    "message"    : f"Merged {len(self.files)} files — {len(merged)} total rows"
                })

        except Exception as e:
            self.error.emit(str(e))