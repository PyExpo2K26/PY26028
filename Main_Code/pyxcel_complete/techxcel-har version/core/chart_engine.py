"""
PyXcel — Chart Engine
Generates charts from Excel data using matplotlib.
"""
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import io
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

COLORS = ["#7c83ff","#4caf81","#ff9800","#f44336","#00bcd4","#e91e63","#9c27b0","#ffeb3b"]
BG     = "#0f1117"
AX_BG  = "#1a1d2e"
TEXT   = "#e0e0e0"
GRID   = "#2a2d3e"


def _style(fig, ax):
    fig.patch.set_facecolor(BG)
    ax.set_facecolor(AX_BG)
    ax.tick_params(colors=TEXT, labelsize=9)
    ax.xaxis.label.set_color(TEXT)
    ax.yaxis.label.set_color(TEXT)
    ax.title.set_color(TEXT)
    for spine in ax.spines.values():
        spine.set_edgecolor(GRID)
    ax.grid(True, color=GRID, linestyle="--", alpha=0.5)


def generate_chart(filepath, sheet, chart_type, x_col, y_col, title="", output_sheet="Charts"):
    try:
        df = pd.read_excel(filepath, sheet_name=sheet)
        if x_col not in df.columns:
            return {"status": "error", "message": f"Column '{x_col}' not found"}
        if y_col not in df.columns:
            return {"status": "error", "message": f"Column '{y_col}' not found"}

        if chart_type in ["bar","line","area"]:
            plot_df = df.groupby(x_col)[y_col].sum().reset_index().sort_values(y_col, ascending=False).head(20)
        else:
            plot_df = df[[x_col, y_col]].dropna().head(100)

        fig, ax = plt.subplots(figsize=(12, 6))

        if chart_type == "bar":
            ax.bar(plot_df[x_col].astype(str), plot_df[y_col], color=COLORS[:len(plot_df)])
            ax.set_xlabel(x_col); ax.set_ylabel(y_col)
            plt.xticks(rotation=45, ha="right")
        elif chart_type == "line":
            ax.plot(plot_df[x_col].astype(str), plot_df[y_col], color=COLORS[0], linewidth=2.5, marker="o", markersize=5)
            ax.fill_between(range(len(plot_df)), plot_df[y_col], alpha=0.1, color=COLORS[0])
            ax.set_xlabel(x_col); ax.set_ylabel(y_col)
            plt.xticks(rotation=45, ha="right")
        elif chart_type == "pie":
            top = plot_df.nlargest(8, y_col)
            wedges, texts, autotexts = ax.pie(top[y_col], labels=top[x_col].astype(str), autopct="%1.1f%%", colors=COLORS[:len(top)], startangle=90)
            for t in texts + autotexts:
                t.set_color(TEXT); t.set_fontsize(9)
        elif chart_type == "scatter":
            ax.scatter(plot_df[x_col], plot_df[y_col], color=COLORS[0], alpha=0.7, s=60)
            ax.set_xlabel(x_col); ax.set_ylabel(y_col)
        elif chart_type == "area":
            ax.fill_between(range(len(plot_df)), plot_df[y_col], alpha=0.4, color=COLORS[0])
            ax.plot(range(len(plot_df)), plot_df[y_col], color=COLORS[0], linewidth=2)
            ax.set_xticks(range(len(plot_df)))
            ax.set_xticklabels(plot_df[x_col].astype(str), rotation=45, ha="right")
        elif chart_type == "histogram":
            numeric_data = pd.to_numeric(df[y_col], errors="coerce").dropna()
            ax.hist(numeric_data, bins=20, color=COLORS[0], edgecolor=GRID, alpha=0.8)
            ax.set_xlabel(y_col); ax.set_ylabel("Frequency")

        ax.set_title(title or f"{chart_type.title()} — {y_col} by {x_col}", fontsize=13, fontweight="bold", pad=15)
        _style(fig, ax)
        plt.tight_layout()

        buf = io.BytesIO()
        plt.savefig(buf, format="png", dpi=150, bbox_inches="tight", facecolor=BG)
        plt.close(fig)
        buf.seek(0)

        wb = load_workbook(filepath)
        if output_sheet not in wb.sheetnames:
            wb.create_sheet(output_sheet)
        ws = wb[output_sheet]

        img        = XLImage(buf)
        img.width  = 700
        img.height = 380
        next_row   = ws.max_row + 3 if ws.max_row > 1 else 1
        ws.add_image(img, f"A{next_row}")
        wb.save(filepath)
        wb.close()

        return {"status": "success", "chart_type": chart_type, "output_sheet": output_sheet, "message": f"{chart_type.title()} chart created in '{output_sheet}'"}
    except Exception as e:
        return {"status": "error", "message": str(e)}


def get_chart_preview(filepath, sheet, chart_type, x_col, y_col, title=""):
    try:
        df = pd.read_excel(filepath, sheet_name=sheet)
        if chart_type in ["bar","line","area"]:
            plot_df = df.groupby(x_col)[y_col].sum().reset_index().sort_values(y_col, ascending=False).head(15)
        else:
            plot_df = df[[x_col, y_col]].dropna().head(100)

        fig, ax = plt.subplots(figsize=(10, 5))

        if chart_type == "bar":
            ax.bar(plot_df[x_col].astype(str), plot_df[y_col], color=COLORS[:len(plot_df)])
            plt.xticks(rotation=45, ha="right")
        elif chart_type == "line":
            ax.plot(plot_df[x_col].astype(str), plot_df[y_col], color=COLORS[0], linewidth=2, marker="o")
            plt.xticks(rotation=45, ha="right")
        elif chart_type == "pie":
            top = plot_df.nlargest(6, y_col)
            ax.pie(top[y_col], labels=top[x_col].astype(str), autopct="%1.1f%%", colors=COLORS[:len(top)])
        elif chart_type == "scatter":
            ax.scatter(plot_df[x_col], plot_df[y_col], color=COLORS[0])
        elif chart_type == "histogram":
            numeric_data = pd.to_numeric(df[y_col], errors="coerce").dropna()
            ax.hist(numeric_data, bins=20, color=COLORS[0])
        elif chart_type == "area":
            ax.fill_between(range(len(plot_df)), plot_df[y_col], alpha=0.4, color=COLORS[0])
            ax.plot(range(len(plot_df)), plot_df[y_col], color=COLORS[0], linewidth=2)

        ax.set_title(title or f"{y_col} by {x_col}", fontsize=12)
        _style(fig, ax)
        plt.tight_layout()

        buf = io.BytesIO()
        plt.savefig(buf, format="png", dpi=120, bbox_inches="tight", facecolor=BG)
        plt.close(fig)
        buf.seek(0)
        return buf.read()
    except Exception:
        return None
