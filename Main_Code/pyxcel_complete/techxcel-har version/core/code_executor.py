"""
Code Executor — Safely runs AI-generated Python/openpyxl/pandas code.
"""
import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import re


def clean_code(raw: str) -> str:
    return re.sub(r"```python|```", "", raw).strip()


def execute_cleaning_code(filepath: str, sheet: str, code: str) -> dict:
    code       = clean_code(code)
    df         = pd.read_excel(filepath, sheet_name=sheet)
    orig_shape = df.shape
    local      = {"df": df, "pd": pd}
    try:
        exec(code, local)
        df = local.get("df", df)
        wb = openpyxl.load_workbook(filepath)
        if sheet in wb.sheetnames:
            del wb[sheet]
        ws = wb.create_sheet(sheet)
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)
        wb.save(filepath)
        wb.close()
        return {"status": "success", "code": code, "original_shape": orig_shape, "new_shape": df.shape, "message": f"Cleaned: {orig_shape[0]} -> {df.shape[0]} rows"}
    except Exception as e:
        return {"status": "error", "code": code, "message": str(e)}
