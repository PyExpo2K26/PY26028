"""
Workbook Inspector — Gives LLaMA live awareness of the Excel file structure.
"""
import openpyxl
import pandas as pd


def inspect_workbook(filepath: str) -> dict:
    wb = openpyxl.load_workbook(filepath, data_only=True)
    summary = {}
    for name in wb.sheetnames:
        ws   = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(h) if h is not None else f"Column_{i}" for i, h in enumerate(rows[0])]
        summary[name] = {
            "headers"  : headers,
            "row_count": ws.max_row - 1,
            "col_count": ws.max_column,
            "sample"   : [dict(zip(headers, row)) for row in rows[1:6]]
        }
    wb.close()
    return summary


def get_context_string(filepath: str) -> str:
    try:
        info  = inspect_workbook(filepath)
        lines = ["=== WORKBOOK STRUCTURE ===\n"]
        for sheet, data in info.items():
            lines.append(f"Sheet: '{sheet}'")
            lines.append(f"  Columns ({data['col_count']}): {', '.join(data['headers'])}")
            lines.append(f"  Total rows: {data['row_count']}")
            lines.append("  Sample data (first 5 rows):")
            for i, row in enumerate(data["sample"], 1):
                lines.append(f"    Row {i}: {row}")
            lines.append("")
        return "\n".join(lines)
    except Exception as e:
        return f"Could not inspect workbook: {str(e)}"


def get_sheet_names(filepath: str) -> list:
    try:
        wb    = openpyxl.load_workbook(filepath, read_only=True)
        names = wb.sheetnames
        wb.close()
        return names
    except Exception:
        return []


def get_dataframe(filepath: str, sheet: str) -> pd.DataFrame:
    return pd.read_excel(filepath, sheet_name=sheet)


def get_column_map(filepath: str, sheet: str) -> dict:
    """Returns exact column numbers for every header."""
    try:
        wb  = openpyxl.load_workbook(filepath, data_only=True)
        ws  = wb[sheet]
        col_map = {}
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value:
                col_map[str(cell.value)] = col_idx
        wb.close()
        return col_map
    except Exception:
        return {}
